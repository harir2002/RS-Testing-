import streamlit as st
import pandas as pd
import openpyxl
import io
import re # Using regex for robust whitespace removal
from datetime import datetime

# --- Helper functions ---
def get_excel_format_type(number_format):
    """Maps Excel's built-in cell format to a more specific conceptual type."""
    if not number_format:
        return "General"
    fmt = number_format.lower()
    if '_(' in fmt and '*' in fmt and ')' in fmt: return 'Accounting'
    if 'yy' in fmt or 'mm' in fmt or 'dd' in fmt: return 'Date'
    if '%' in fmt: return 'Percentage'
    if any(c in fmt for c in ['$', '€', '£', '¥', '₹']): return 'Currency'
    if '0' in fmt or '#' in fmt: return 'Numeric'
    if fmt == '@': return 'Text'
    if fmt == 'general': return 'General'
    return "Other"

def normalize_value_for_comparison(value):
    """
    Intelligently normalizes a value to its true content for comparison.
    Handles numbers stored as text and removes ALL types of whitespace.
    """
    if isinstance(value, str):
        cleaned_value = value.strip().lower()
        cleaned_value = re.sub(r'\s+', '', cleaned_value)
        try:
            if '.' in cleaned_value:
                return float(cleaned_value)
            return int(cleaned_value)
        except (ValueError, TypeError):
            return cleaned_value
    return value

# ---------------- Core Validation Logic with Simplified Reason Text ----------------
def compare_excel_files(input_file, output_file):
    results = {}
    try:
        input_wb_fmt = openpyxl.load_workbook(input_file, data_only=False)
        output_wb_fmt = openpyxl.load_workbook(output_file, data_only=False)
        input_wb_val = openpyxl.load_workbook(input_file, data_only=True)
        output_wb_val = openpyxl.load_workbook(output_file, data_only=True)
    except Exception as e:
        st.error(f"Error: Could not load Excel files. Details: {e}")
        return {}

    sheets_to_process = input_wb_val.sheetnames
    
    for sheet_name in sheets_to_process:
        try:
            results[sheet_name] = []
            ws_in_val = input_wb_val[sheet_name]
            num_cols, num_rows = ws_in_val.max_column, ws_in_val.max_row
            headers = {c: ws_in_val.cell(row=3, column=c).value for c in range(1, num_cols + 1)}

            if sheet_name not in output_wb_val.sheetnames:
                st.warning(f"Sheet '{sheet_name}' from the template is MISSING in the output file.")
                for c in range(1, num_cols + 1):
                    for r in range(3, num_rows + 1):
                        t_val = ws_in_val.cell(row=r, column=c).value
                        if t_val is None or str(t_val).strip() == '': continue
                        results[sheet_name].append({
                            "SHEET": sheet_name, "CELL": ws_in_val.cell(row=r, column=c).coordinate,
                            "FIELD": headers.get(c, f"Col_{c}"), "EXPECTED VALUE": str(t_val),
                            "TEST VALUE": "N/A (Sheet Missing)", "Data Type Result": "Wrong",
                            "Data Type Reason": f"Sheet '{sheet_name}' not in output.",
                            "Value Result": "Wrong", "Value Reason": f"Sheet '{sheet_name}' not in output."
                        })
                continue

            ws_in_fmt = input_wb_fmt[sheet_name]
            ws_out_fmt = output_wb_fmt[sheet_name]
            ws_out_val = output_wb_val[sheet_name]

            # --- THIS IS THE CHANGE: Simplified reason function ---
            def get_mismatch_reason(template_val, output_val):
                return f"The template value is `{template_val}`, but the output has `{output_val}`."

            for c in range(1, num_cols + 1):
                for r in range(3, num_rows + 1):
                    t_val, o_val = ws_in_val.cell(row=r, column=c).value, ws_out_val.cell(row=r, column=c).value
                    if t_val is None and o_val is None: continue
                    if t_val is None or str(t_val).strip() == '': continue

                    t_fmt, o_fmt = ws_in_fmt.cell(row=r, column=c).number_format, ws_out_fmt.cell(row=r, column=c).number_format
                    t_type, o_type = get_excel_format_type(t_fmt), get_excel_format_type(o_fmt)
                    is_dtype_match = t_type == o_type
                    dtype_res = "Correct" if is_dtype_match else "Wrong"
                    dtype_reason = "Data types match" if is_dtype_match else f"Template type is `{t_type}`, but output is `{o_type}`."
                    
                    is_match = normalize_value_for_comparison(t_val) == normalize_value_for_comparison(o_val)
                    val_res = "Correct" if is_match else "Wrong"
                    val_reason = "Values match" if is_match else get_mismatch_reason(t_val, o_val)

                    if r == 3:
                        dtype_res, dtype_reason = "N/A", "Header row - no type check"

                    results[sheet_name].append({
                        "SHEET": sheet_name, "CELL": ws_out_val.cell(row=r, column=c).coordinate,
                        "FIELD": headers.get(c, f"Col_{c}"), "EXPECTED VALUE": str(t_val),
                        "TEST VALUE": str(o_val), "Data Type Result": dtype_res,
                        "Data Type Reason": dtype_reason, "Value Result": val_res,
                        "Value Reason": val_reason
                    })
        except Exception as e:
            st.warning(f"Error processing sheet '{sheet_name}': {e}")
            continue
    return results

# ---------------- Report Generation and UI (No changes below this line) ----------------
def generate_excel_report(results):
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    workbook = writer.book
    all_rows = [row for sheet_rows in results.values() for row in sheet_rows]
    if not all_rows: return None
    df = pd.DataFrame(all_rows)
    
    total_value_checks = len(df)
    data_df = df[df["Data Type Result"] != "N/A"].copy()
    total_dtype_checks = len(data_df)
    
    dtype_correct = len(data_df[data_df["Data Type Result"] == "Correct"])
    value_correct = len(df[df["Value Result"] == "Correct"])
    
    dtype_accuracy = (dtype_correct / total_dtype_checks * 100) if total_dtype_checks > 0 else 100
    value_accuracy = (value_correct / total_value_checks * 100) if total_value_checks > 0 else 100
    
    dtype_errors = total_dtype_checks - dtype_correct
    value_errors = total_value_checks - value_correct

    dash = workbook.add_worksheet("QA Dashboard")
    title_fmt, kpi_fmt, label_fmt = workbook.add_format({'bold': True, 'font_size': 20, 'align': 'center', 'valign': 'vcenter'}), workbook.add_format({'bold': True, 'font_size': 28, 'align': 'center', 'valign': 'vcenter'}), workbook.add_format({'font_size': 12, 'align': 'center', 'font_color': '#595959'})
    dash.merge_range('B2:G3', 'Validation Dashboard', title_fmt); dash.merge_range('B5:D7', f"{dtype_accuracy:.1f}%", kpi_fmt); dash.merge_range('B8:D8', 'Data Type Accuracy', label_fmt); dash.merge_range('E5:G7', f"{value_accuracy:.1f}%", kpi_fmt); dash.merge_range('E8:G8', 'Value Accuracy', label_fmt); dash.merge_range('B10:D12', dtype_errors, kpi_fmt); dash.merge_range('B13:D13', 'Data Type Errors', label_fmt); dash.merge_range('E10:G12', value_errors, kpi_fmt); dash.merge_range('E13:G13', 'Value Errors', label_fmt); dash.set_column('B:G', 22)
    cell_wrap_format = workbook.add_format({'text_wrap': True, 'valign': 'top'})
    common_props = {'bold': True, 'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'font_color': '#FFFFFF'}
    header_fmt_blue, header_fmt_red, header_fmt_green = workbook.add_format({**common_props, 'bg_color': '#002060'}), workbook.add_format({**common_props, 'bg_color': '#C00000'}), workbook.add_format({**common_props, 'bg_color': '#00B050'})
    dtype_cols = ["SHEET", "CELL", "FIELD", "Data Type Result", "Data Type Reason"]
    dtype_df_to_excel = df[df["Data Type Result"] != "N/A"][dtype_cols] if "Data Type Result" in df else pd.DataFrame(columns=dtype_cols)
    dtype_df_to_excel.to_excel(writer, sheet_name="Data Type Results", index=False, header=False, startrow=1)
    ws_dtype = writer.sheets["Data Type Results"]
    for col_num, value in enumerate(dtype_df_to_excel.columns.values):
        ws_dtype.write(0, col_num, value, header_fmt_blue if value in ["SHEET", "CELL", "FIELD"] else header_fmt_green)
    ws_dtype.set_column('A:E', 25, cell_wrap_format)
    value_cols = ["SHEET", "CELL", "FIELD", "EXPECTED VALUE", "TEST VALUE", "Value Result", "Value Reason"]
    value_df = df[value_cols] if value_cols[0] in df else pd.DataFrame(columns=value_cols)
    value_df.to_excel(writer, sheet_name="Value Match Results", index=False, header=False, startrow=1)
    ws_value = writer.sheets["Value Match Results"]
    for col_num, value in enumerate(value_df.columns.values):
        if value in ["SHEET", "CELL", "FIELD"]: header_format = header_fmt_blue
        elif value in ["EXPECTED VALUE", "TEST VALUE"]: header_format = header_fmt_red
        else: header_format = header_fmt_green
        ws_value.write(0, col_num, value, header_format)
    ws_value.set_column('A:G', 25, cell_wrap_format)
    writer.close()
    output.seek(0)
    return output

st.set_page_config(page_title="Excel Validator", layout="wide")
st.title("Excel Validator — Data Type & Value Check")
if 'ran' not in st.session_state: st.session_state.ran = False
if 'results' not in st.session_state: st.session_state.results = {}
input_file = st.file_uploader("Upload Input Template Excel", type=['xlsx'])
output_file = st.file_uploader("Upload Output Excel to Test", type=['xlsx'])
if input_file and output_file:
    if st.button("Run Validation", type="primary"):
        with st.spinner("Validating..."):
            st.session_state.results = compare_excel_files(input_file, output_file)
        st.session_state.ran = True
if st.session_state.ran:
    res = st.session_state.results
    if res:
        report_data = generate_excel_report(res)
        if report_data:
            all_rows = [row for sheet_rows in res.values() for row in sheet_rows]
            df = pd.DataFrame(all_rows)
            total_value_checks = len(df)
            data_df = df[df["Data Type Result"] != "N/A"].copy()
            total_dtype_checks = len(data_df)
            
            dtype_correct = len(data_df[data_df["Data Type Result"] == "Correct"])
            value_correct = len(df[df["Value Result"] == "Correct"])
            
            dtype_accuracy = (dtype_correct / total_dtype_checks * 100) if total_dtype_checks > 0 else 100
            value_accuracy = (value_correct / total_value_checks * 100) if total_value_checks > 0 else 100
            
            dtype_errors = total_dtype_checks - dtype_correct
            value_errors = total_value_checks - value_correct
            
            st.header("Validation Summary")
            col1, col2 = st.columns(2)
            col1.metric("📊 Data Type Accuracy", f"{dtype_accuracy:.1f}%", f"{dtype_errors} Errors", delta_color="inverse")
            col2.metric("🔢 Value Accuracy", f"{value_accuracy:.1f}%", f"{value_errors} Errors", delta_color="inverse")
            st.download_button("📄 Download Full Test Report", data=report_data, file_name=f"Test_Report_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", on_click=lambda: st.session_state.clear())
